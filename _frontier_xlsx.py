import sys, io, pickle, numpy as np, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

with open('product_paths.pkl','rb') as f:
    data = pickle.load(f)

W0=100.0

def sim(path, wr, mode, bl=0.05, bh=0.05, thr=1.0):
    tr=wr/12;nav=W0;cum=0;prev_wd=W0*tr;wds=[]
    for t,r in enumerate(path):
        if mode=='vol':
            if t>=12:
                s=np.std(path[t-12:t],ddof=1)
                sr=abs(path[t-1])/s if s>0 else 0
                band=bh if sr>=thr else bl
            else: band=bl
        else: band=bl
        wd=prev_wd
        if nav>0:
            u=tr*(1+band);l=tr*(1-band);ratio=wd/nav
            if ratio>u:wd=u*nav
            elif ratio<l:wd=l*nav
        wd=min(wd,max(nav,0));nav=max(nav-wd,0)*(1+r);prev_wd=wd;cum+=wd;wds.append(wd)
    ws=np.array(wds);wp=ws[ws>0]
    cuts=[(ws[i]-ws[i-1])/ws[i-1] for i in range(1,len(ws)) if ws[i-1]>0]
    return {'nav':nav,'cum':cum,'tot':nav+cum,'std':np.std(wp) if len(wp)>0 else 0,'worst':min(cuts) if cuts else 0}

def med(results,key):
    return np.median([r[key] for r in results])

bands=[0.001,0.005,0.01,0.015,0.02,0.025,0.03,0.035,0.04,0.045,0.05,
       0.06,0.07,0.08,0.09,0.10,0.12,0.15,0.18,0.20,0.25,0.30,0.40,0.50,0.75,1.0,2.0,5.0]
vol_targets = [(0.05, 0.08), (0.05, 0.10), (0.05, 0.15)]
flist=['Golden Growth','MS GROWTH','MS STABLE']

hdr_font = Font(name='Pretendard', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='1a237e')
data_font = Font(name='Pretendard', size=10)
bold_font = Font(name='Pretendard', size=10, bold=True)
alpha_font = Font(name='Pretendard', size=10, bold=True, color='FF6F00')
g5_fill = PatternFill('solid', fgColor='FFEBEE')
alpha_fill = PatternFill('solid', fgColor='FFF3E0')
center = Alignment(horizontal='center')

wb = openpyxl.Workbook()
wb.remove(wb.active)

for fund in flist:
    mr = data['fund_paths'][fund]['monthly_returns']
    wr = ((1+mr.mean())**12-1)+0.01
    paths = data['fund_paths'][fund]['bootstrap_paths']
    tr = wr/12

    # ============================================================
    # Compute
    # ============================================================
    bd_data=[]
    for band in bands:
        res=[sim(p,wr,'fixed',band) for p in paths]
        d={k:med(res,k) for k in ['nav','cum','tot','std','worst']}
        d['worst_abs']=abs(d['worst'])*100
        d['band']=band
        bd_data.append(d)

    fa_res=[sim(p,wr,'fixed',99.0) for p in paths]
    fa={k:med(fa_res,k) for k in ['nav','cum','tot','std','worst']}
    fa['worst_abs']=abs(fa['worst'])*100

    fr_list=[]
    for p in paths:
        n=W0;c=0;wds=[]
        for r in p: wd=n*tr;n=max(n-wd,0)*(1+r);c+=wd;wds.append(wd)
        ws_arr=np.array(wds);wp=ws_arr[ws_arr>0]
        cuts=[(ws_arr[i]-ws_arr[i-1])/ws_arr[i-1] for i in range(1,len(ws_arr)) if ws_arr[i-1]>0]
        fr_list.append({'nav':n,'cum':c,'tot':n+c,
                        'std':np.std(wp) if len(wp)>0 else 0,
                        'worst':min(cuts) if cuts else 0})
    fr={k:np.median([r[k] for r in fr_list]) for k in ['nav','cum','tot','std','worst']}
    fr['worst_abs']=abs(fr['worst'])*100

    g5_res=[sim(p,wr,'fixed',0.05) for p in paths]
    g5={k:med(g5_res,k) for k in ['nav','cum','tot','std','worst']}
    g5['worst_abs']=abs(g5['worst'])*100

    vol_data=[]
    for bl,bh in vol_targets:
        res=[sim(p,wr,'vol',bl,bh) for p in paths]
        d={k:med(res,k) for k in ['nav','cum','tot','std','worst']}
        d['worst_abs']=abs(d['worst'])*100
        d['bl']=bl;d['bh']=bh
        d['label']=f'{bl*100:.0f}/{bh*100:.0f}%'
        vol_data.append(d)

    # ============================================================
    # Sheet: Tot/Worst
    # ============================================================
    sn = fund.replace(' ','')[:15] + '_TotWorst'
    sheet = wb.create_sheet(sn[:31])

    sheet.merge_cells('A1:H1')
    sheet['A1'] = f'{fund} — Efficient Frontier: Total vs Worst Cut (WR={wr*100:.1f}%)'
    sheet['A1'].font = Font(name='Pretendard', bold=True, size=13)

    # Headers
    headers = ['Series','Label','Band','NAV','Cum','Total','Worst Cut(%)','Std']
    for c, h in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center

    def write_row(sheet, row, series, label, nav, cum, tot, worst_abs, std, fill=None, font=None):
        vals = [series, label, '', round(nav,2), round(cum,2), round(tot,2), round(worst_abs,2), round(std,4)]
        for c, v in enumerate(vals, 1):
            cell = sheet.cell(row=row, column=c, value=v)
            cell.font = font or data_font
            if fill: cell.fill = fill

    row = 4
    for d in bd_data:
        vals = ['Frontier', f"{d['band']*100:.1f}%", d['band'],
                round(d['nav'],2), round(d['cum'],2), round(d['tot'],2),
                round(d['worst_abs'],2), round(d['std'],4)]
        for c, v in enumerate(vals, 1):
            sheet.cell(row=row, column=c, value=v).font = data_font
        row += 1

    # FA
    vals = ['Fixed Amt', 'FA', 99.0, round(fa['nav'],2), round(fa['cum'],2),
            round(fa['tot'],2), round(fa['worst_abs'],2), round(fa['std'],4)]
    for c, v in enumerate(vals, 1):
        sheet.cell(row=row, column=c, value=v).font = data_font
    row += 1

    # FR
    vals = ['Fixed Rate', 'FR', 0, round(fr['nav'],2), round(fr['cum'],2),
            round(fr['tot'],2), round(fr['worst_abs'],2), round(fr['std'],4)]
    for c, v in enumerate(vals, 1):
        sheet.cell(row=row, column=c, value=v).font = data_font
    row += 1

    # G5%
    vals = ['Guard 5%', 'G5%', 0.05, round(g5['nav'],2), round(g5['cum'],2),
            round(g5['tot'],2), round(g5['worst_abs'],2), round(g5['std'],4)]
    for c, v in enumerate(vals, 1):
        cell = sheet.cell(row=row, column=c, value=v)
        cell.font = bold_font; cell.fill = g5_fill
    row += 1

    # Vol combos
    for d in vol_data:
        vals = ['Vol-Adj', d['label'], f"{d['bl']}/{d['bh']}",
                round(d['nav'],2), round(d['cum'],2), round(d['tot'],2),
                round(d['worst_abs'],2), round(d['std'],4)]
        for c, v in enumerate(vals, 1):
            cell = sheet.cell(row=row, column=c, value=v)
            cell.font = alpha_font; cell.fill = alpha_fill
        row += 1

    for c,w in [(1,12),(2,10),(3,8),(4,10),(5,10),(6,10),(7,14),(8,10)]:
        sheet.column_dimensions[get_column_letter(c)].width = w

    # ============================================================
    # Sheet: Tot/Std
    # ============================================================
    sn2 = fund.replace(' ','')[:15] + '_TotStd'
    sheet2 = wb.create_sheet(sn2[:31])

    sheet2.merge_cells('A1:H1')
    sheet2['A1'] = f'{fund} — Efficient Frontier: Total vs Std (WR={wr*100:.1f}%)'
    sheet2['A1'].font = Font(name='Pretendard', bold=True, size=13)

    headers2 = ['Series','Label','Band','NAV','Cum','Total','Std','Worst Cut(%)']
    for c, h in enumerate(headers2, 1):
        cell = sheet2.cell(row=3, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center

    row = 4
    for d in bd_data:
        vals = ['Frontier', f"{d['band']*100:.1f}%", d['band'],
                round(d['nav'],2), round(d['cum'],2), round(d['tot'],2),
                round(d['std'],4), round(d['worst_abs'],2)]
        for c, v in enumerate(vals, 1):
            sheet2.cell(row=row, column=c, value=v).font = data_font
        row += 1

    # FA
    vals = ['Fixed Amt', 'FA', 99.0, round(fa['nav'],2), round(fa['cum'],2),
            round(fa['tot'],2), round(fa['std'],4), round(fa['worst_abs'],2)]
    for c, v in enumerate(vals, 1):
        sheet2.cell(row=row, column=c, value=v).font = data_font
    row += 1

    # FR
    vals = ['Fixed Rate', 'FR', 0, round(fr['nav'],2), round(fr['cum'],2),
            round(fr['tot'],2), round(fr['std'],4), round(fr['worst_abs'],2)]
    for c, v in enumerate(vals, 1):
        sheet2.cell(row=row, column=c, value=v).font = data_font
    row += 1

    # G5%
    vals = ['Guard 5%', 'G5%', 0.05, round(g5['nav'],2), round(g5['cum'],2),
            round(g5['tot'],2), round(g5['std'],4), round(g5['worst_abs'],2)]
    for c, v in enumerate(vals, 1):
        cell = sheet2.cell(row=row, column=c, value=v)
        cell.font = bold_font; cell.fill = g5_fill
    row += 1

    # Vol combos
    for d in vol_data:
        vals = ['Vol-Adj', d['label'], f"{d['bl']}/{d['bh']}",
                round(d['nav'],2), round(d['cum'],2), round(d['tot'],2),
                round(d['std'],4), round(d['worst_abs'],2)]
        for c, v in enumerate(vals, 1):
            cell = sheet2.cell(row=row, column=c, value=v)
            cell.font = alpha_font; cell.fill = alpha_fill
        row += 1

    for c,w in [(1,12),(2,10),(3,8),(4,10),(5,10),(6,10),(7,10),(8,14)]:
        sheet2.column_dimensions[get_column_letter(c)].width = w

wb.save('frontier_data.xlsx')
print(f'frontier_data.xlsx saved ({len(wb.sheetnames)} sheets: {wb.sheetnames})')
